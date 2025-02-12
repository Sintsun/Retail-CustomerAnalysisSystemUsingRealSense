import argparse
import logging
import matplotlib
matplotlib.use('TkAgg')
import pyrealsense2 as rs
import numpy as np
import cv2
import os
import mediapipe as mp
import time
import datetime
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
import matplotlib.patches as mpatches
from openpyxl import Workbook, load_workbook
import queue
import threading
from collections import deque

##########################################################
# Global Crop Size Settings
##########################################################
CROP_WIDTH = 640   # Fixed crop width
CROP_HEIGHT = 480  # Fixed crop height

##########################################################
# (A) Define a log handler that stores recent log messages in a deque
##########################################################
MAX_LOG_LINES = 20  # Maximum number of log lines to store
log_deque = deque(maxlen=MAX_LOG_LINES)

class Cv2LogHandler(logging.Handler):
    def emit(self, record):
        msg = self.format(record)
        log_deque.append(msg)  # Append each log message to the deque

# Configure basic logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - [%(levelname)s] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
cv2_log_handler = Cv2LogHandler()
cv2_log_handler.setLevel(logging.INFO)
cv2_log_handler.setFormatter(
    logging.Formatter('%(asctime)s - [%(levelname)s] - %(message)s',
                      datefmt='%Y-%m-%d %H:%M:%S')
)
logging.getLogger().addHandler(cv2_log_handler)

##########################################################
# (B) Excel Writing Functions and Thread for Periodic Flushing
##########################################################
FLUSH_INTERVAL = 2.0      # Check every 2 seconds whether to flush logs to Excel
FLUSH_BATCH_SIZE = 10     # Flush logs when there are 10 or more records in the buffer

excel_queue = queue.Queue()
flush_event = threading.Event()
stop_event = threading.Event()

def log_intersection_2d_async(face_id, plane_name, u, v, timestamp):
    """
    Enqueue a log record for later writing to Excel.
    """
    record = (timestamp, face_id, plane_name, u, v)
    excel_queue.put(record)
    if excel_queue.qsize() >= FLUSH_BATCH_SIZE:
        flush_event.set()

def flush_excel_queue(xlsx_path):
    """
    Flush all records in the excel_queue to the Excel file.
    """
    temp_records = []
    while not excel_queue.empty():
        temp_records.append(excel_queue.get())
    if len(temp_records) == 0:
        return
    headers = ["Timestamp", "Face_ID", "Plane_Name", "u_local", "v_local"]
    file_exists = os.path.exists(xlsx_path)
    try:
        if file_exists:
            wb = load_workbook(xlsx_path, keep_vba=True)
            ws = wb.active
            if ws.max_row == 0:
                ws.append(headers)
            else:
                if ws['A1'].value != headers[0]:
                    ws.insert_rows(1)
                    for col_idx, val in enumerate(headers, start=1):
                        ws.cell(row=1, column=col_idx, value=val)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
        for (ts_, face_id_, plane_name_, u_, v_) in temp_records:
            ws.append([ts_, face_id_, plane_name_, u_, v_])
        wb.save(xlsx_path)
        logging.info(f"[Excel] Wrote {len(temp_records)} records to {xlsx_path}")
    except Exception as e:
        logging.error(f"Failed to write to Excel {xlsx_path}, error: {e}")

def flush_thread_func(xlsx_path):
    """
    Thread function that periodically flushes the Excel queue.
    """
    while not stop_event.is_set():
        triggered = flush_event.wait(timeout=FLUSH_INTERVAL)
        flush_excel_queue(xlsx_path)
        flush_event.clear()
    flush_excel_queue(xlsx_path)

def get_max_face_id_from_xlsm(xlsx_path):
    """
    Retrieve the maximum Face_ID from the Excel log file.
    """
    if not os.path.exists(xlsx_path):
        logging.warning(f"Excel file not found: {xlsx_path}")
        return -1
    try:
        wb = load_workbook(xlsx_path, keep_vba=True)
        ws = wb.active
        if ws.max_row < 2:
            wb.close()
            logging.warning("No valid data row in Excel (only header or empty).")
            return -1
        header_row = [cell.value for cell in ws[1]]
        if "Face_ID" not in header_row:
            wb.close()
            logging.warning("Column 'Face_ID' not found in Excel.")
            return -1
        face_id_col_idx = header_row.index("Face_ID") + 1
        max_id = -1
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=face_id_col_idx).value
            if val is not None:
                try:
                    face_id_int = int(val)
                    if face_id_int > max_id:
                        max_id = face_id_int
                except ValueError:
                    logging.warning(f"Row {row_idx} Face_ID is not an integer: {val}")
        wb.close()
        return max_id
    except Exception as e:
        logging.error(f"Cannot read {xlsx_path}, error: {e}")
        return -1

##########################################################
# Dummy Implementation for ransac_weighted_kabsch
##########################################################
def ransac_weighted_kabsch(ref_points, det_points, weights, threshold, max_iterations):
    """
    A dummy implementation of the weighted Kabsch algorithm with RANSAC.
    It returns a rotation matrix R and translation vector t.
    """
    total_weight = np.sum(weights)
    ref_centroid = np.sum(ref_points * weights, axis=1, keepdims=True) / total_weight
    det_centroid = np.sum(det_points * weights, axis=1, keepdims=True) / total_weight
    ref_centered = ref_points - ref_centroid
    det_centered = det_points - det_centroid
    W = np.diag(weights)
    covariance_matrix = ref_centered @ W @ det_centered.T
    U, _, Vt = np.linalg.svd(covariance_matrix)
    R = Vt.T @ U.T
    if np.linalg.det(R) < 0:
        Vt[-1, :] *= -1
        R = Vt.T @ U.T
    t = det_centroid - R @ ref_centroid
    inliers = np.ones(det_points.shape[1], dtype=bool)
    return (R, t), inliers

##########################################################
# KalmanBoxTracker and SORT Tracker Classes
##########################################################
class KalmanBoxTracker:
    count = 0

    def __init__(self, bbox):
        self.bbox = bbox  # Bounding box [x1, y1, x2, y2]
        self.id = KalmanBoxTracker.count
        KalmanBoxTracker.count += 1
        self.hits = 1
        self.age = 0
        self.time_since_update = 0

    def update(self, bbox):
        if bbox is not None:
            self.bbox = bbox
            self.hits += 1
            self.time_since_update = 0
        else:
            self.time_since_update += 1
        self.age += 1

    def get_state(self):
        return self.bbox


class Sort:
    def __init__(self, max_age=10, min_hits=3):
        self.max_age = max_age  # Maximum frames to keep a tracker without update
        self.min_hits = min_hits  # Minimum hits to confirm a tracker
        self.trackers = []

    def update(self, dets):
        """
        Update trackers with new detections.
        """
        matched, unmatched_dets, unmatched_trks = self.associate_detections_to_trackers(dets, self.trackers)
        for m in matched:
            self.trackers[m[1]].update(dets[m[0]])
        for idx in unmatched_dets:
            self.trackers.append(KalmanBoxTracker(dets[idx]))
        for t_idx in unmatched_trks:
            self.trackers[t_idx].update(None)
        ret = []
        for t in self.trackers[::-1]:
            d = t.get_state()
            if (t.hits >= self.min_hits) and (t.time_since_update < 1):
                ret.append([d[0], d[1], d[2], d[3], t.id])
            if t.time_since_update > self.max_age:
                self.trackers.remove(t)
        return ret

    def associate_detections_to_trackers(self, dets, trackers):
        """
        Associate detection bounding boxes to tracker bounding boxes using IOU.
        """
        if len(trackers) == 0:
            return [], list(range(len(dets))), []
        iou_matrix = np.zeros((len(dets), len(trackers)), dtype=np.float32)
        for d, det in enumerate(dets):
            for t, trk in enumerate(trackers):
                iou_matrix[d, t] = self.iou(det, trk.get_state())
        matched_indices = []
        unmatched_dets = list(range(len(dets)))
        unmatched_trks = list(range(len(trackers)))
        for _ in range(min(len(dets), len(trackers))):
            max_val = np.max(iou_matrix)
            if max_val < 0.1:
                break
            d_ind, t_ind = np.unravel_index(np.argmax(iou_matrix), iou_matrix.shape)
            matched_indices.append([d_ind, t_ind])
            iou_matrix[d_ind, :] = -1
            iou_matrix[:, t_ind] = -1
            unmatched_dets.remove(d_ind)
            unmatched_trks.remove(t_ind)
        return matched_indices, unmatched_dets, unmatched_trks

    def iou(self, bb_test, bb_gt):
        """
        Compute Intersection over Union (IOU) between two bounding boxes.
        """
        xx1 = np.maximum(bb_test[0], bb_gt[0])
        yy1 = np.maximum(bb_test[1], bb_gt[1])
        xx2 = np.minimum(bb_test[2], bb_gt[2])
        yy2 = np.minimum(bb_test[3], bb_gt[3])
        w = np.maximum(0., xx2 - xx1)
        h = np.maximum(0., yy2 - yy1)
        wh = w * h
        o = wh / (
            (bb_test[2] - bb_test[0]) * (bb_test[3] - bb_test[1]) +
            (bb_gt[2] - bb_gt[0]) * (bb_gt[3] - bb_gt[1]) - wh
        )
        return o

##########################################################
# Global Parameters and Settings
##########################################################
XLSX_PATH = r"./log/intersection_log_1.xlsm"
IMAGE_FOLDER = "./data"
RVEC_PATH = "rotation_vectors.npy"
TVEC_PATH = "translation_vectors.npy"
CAMERA_MATRIX_PATH = "rgb_intrinsic_matrix.npy"
DIST_COEFFS_PATH = "dist_coeffs.npy"

FORWARD_LENGTH_3D = 0.8  # Forward length for 3D vector (meters)
FORWARD_LENGTH_2D = 0.2  # Forward length for 2D vector (meters)

# 3D plot axes limits (display in feet)
X_LIMITS = (-10, 10)
Y_LIMITS = (-16, 16)
Z_LIMITS = (-10, 3)

RANSAC_THRESHOLD = 0.01
RANSAC_MAX_ITERATIONS = 150

FRAMERATE = 30.0
FRAME_INTERVAL = 1.0 / FRAMERATE

MIN_DETECTION_CONFIDENCE = 0.5
MIN_TRACKING_CONFIDENCE = 0.5

REFERENCE_LANDMARK_INDICES = [4, 33, 133, 263, 362, 61, 291, 13]
STABLE_POINTS_SET = [4, 33, 133, 263, 362, 61, 291, 13]
SMOOTH_WINDOW_SIZE = 3
INITIALIZATION_FRAMES = 10

# Original cube data is defined in feet. Colors are adjusted to be as different as possible.
CUBES_INFO = [
    {
        "name": "Plane1",
        "color": "cyan",
        "face_points": [
            [4, 7, -5],
            [10, 7, -5],
            [4, 7, -4]
        ],
        "depth_point": [0, 7.5, 0],
        "invert_x": True,
        "invert_y": True,
        "record_face": 0
    },
    {
        "name": "Plane2",
        "color": "magenta",
        "face_points": [
            [-1, 5, -1],
            [4, 5, -1],
            [-1, 5, 3]
        ],
        "depth_point": [0, 6, 0],
        "invert_x": True,
        "invert_y": True,
        "record_face": 0
    },
    {
        "name": "Plane3",
        "color": "yellow",
        "face_points": [
            [-4.5, 5, -2],
            [-3, 5, -2],
            [-4.5, 5, 3]
        ],
        "depth_point": [0, 13, 0],
        "invert_x": True,
        "invert_y": True,
        "record_face": 0
    },
    {
        "name": "Plane4",
        "color": "blue",
        "face_points": [
            [-6, -2.5, -5],
            [-6, 18.5, -5],
            [-6, -2.5, 3]
        ],
        "depth_point": [-6.5, 0, 0],
        "invert_x": True,
        "invert_y": True,
        "record_face": 1
    },
    {
        "name": "Plane5",
        "color": "orange",
        "face_points": [
            [-6, 18.5, -5],
            [8, 18.5, -5],
            [-6, 18.5, 3]
        ],
        "depth_point": [0, 19, 0],
        "invert_x": True,
        "invert_y": True,
        "record_face": 0
    },
    {
        "name": "Plane6",
        "color": "green",
        "face_points": [
            [5.5, 5, -0.5],
            [8, 5, -0.5],
            [5.5, 5, 3]
        ],
        "depth_point": [0, 13, 0],
        "invert_x": True,
        "invert_y": True,
        "record_face": 0
    },
    {
        "name": "Plane7",
        "color": "purple",
        "face_points": [
            [10, -2.5, -5],
            [10, 18, -5],
            [10, -2.5, 3]
        ],
        "depth_point": [10.5, 0, 0],
        "invert_x": True,
        "invert_y": False,
        "record_face": 1
    }
]

##########################################################
# Modified plot_cube_with_index: Returns a list of (face_idx, face_points)
##########################################################
def plot_cube_with_index(ax, face_points, depth_point, color='r', alpha=0.5):
    """
    Given three face points and a depth point, compute all 6 faces of a cube.
    If ax is provided, plot the cube in a 3D matplotlib axis.
    """
    p0, p1, p2 = face_points
    vec1 = p1 - p0
    vec2 = p2 - p0
    if np.allclose(np.cross(vec1, vec2), 0):
        raise ValueError("Invalid face points (collinear).")
    p3 = p1 + (p2 - p0)
    front_face = [p0, p1, p3, p2]
    normal = np.cross(vec1, vec2)
    normal /= np.linalg.norm(normal)
    dp_vec = depth_point - p0
    depth = np.dot(dp_vec, normal)
    offset = depth * normal
    back_face = [v + offset for v in front_face]
    side1 = [front_face[0], front_face[1], back_face[1], back_face[0]]
    side2 = [front_face[1], front_face[2], back_face[2], back_face[1]]
    side3 = [front_face[2], front_face[3], back_face[3], back_face[2]]
    side4 = [front_face[3], front_face[0], back_face[0], back_face[3]]
    cube_faces = []
    cube_faces.append((0, front_face))
    cube_faces.append((1, back_face))
    cube_faces.append((2, side1))
    cube_faces.append((3, side2))
    cube_faces.append((4, side3))
    cube_faces.append((5, side4))
    if ax is not None:
        poly3d_list = [face for idx, face in cube_faces]
        poly = Poly3DCollection(poly3d_list, facecolors=color, edgecolors='k', alpha=alpha)
        ax.add_collection3d(poly)
    return cube_faces

##########################################################
# Additional function: Plot each vertex of a face with different colors
##########################################################
def plot_face_points_with_colors(ax, face, colors=None, s=80):
    """
    Plot each point of the given face with different colors and annotate them.
    """
    if colors is None:
        colors = ['red', 'green', 'blue', 'yellow']
    for i, pt in enumerate(face):
        ax.scatter(pt[0], pt[1], pt[2], color=colors[i % len(colors)], s=s)
        ax.text(pt[0], pt[1], pt[2], f'{i}', color='black', fontsize=10)

##########################################################
# Geometry and Detection Related Functions
##########################################################
def load_distortion_coeffs(dist_coeffs_path):
    """
    Load camera distortion coefficients from a file.
    """
    if not os.path.exists(dist_coeffs_path):
        logging.warning(f"Distortion file not found: {dist_coeffs_path}. Assuming no distortion.")
        return np.zeros((5, 1), dtype=np.float32)
    return np.load(dist_coeffs_path).astype(np.float32)

def load_vectors(rvec_path, tvec_path):
    """
    Load rotation and translation vectors from files.
    """
    if not os.path.exists(rvec_path):
        raise FileNotFoundError(f"Rotation vector not found: {rvec_path}")
    if not os.path.exists(tvec_path):
        raise FileNotFoundError(f"Translation vector not found: {tvec_path}")
    rvec = np.load(rvec_path).astype(np.float32).squeeze()
    tvec = np.load(tvec_path).astype(np.float32).squeeze()
    rvec = rvec.reshape(3, 1) if rvec.shape == (3,) else rvec
    tvec = tvec.reshape(3, 1) if tvec.shape == (3,) else tvec
    return rvec.astype(np.float32), tvec.astype(np.float32)

def initialize_realsense():
    """
    Initialize the RealSense pipeline and return the pipeline, depth scale,
    camera intrinsics, and distortion coefficients.
    """
    logging.info("Initializing RealSense pipeline...")
    pipeline = rs.pipeline()
    config = rs.config()
    config.enable_stream(rs.stream.depth, 1280, 720, rs.format.z16, 30)
    config.enable_stream(rs.stream.color, 1280, 720, rs.format.bgr8, 30)
    profile = pipeline.start(config)
    depth_sensor = profile.get_device().first_depth_sensor()
    depth_scale = depth_sensor.get_depth_scale()
    logging.info(f"Depth scale: {depth_scale} meters/unit.")
    color_stream = profile.get_stream(rs.stream.color).as_video_stream_profile()
    intrinsics = color_stream.get_intrinsics()
    distortion_coeffs = np.array(intrinsics.coeffs, dtype=np.float32).reshape(5, 1)
    return pipeline, depth_scale, intrinsics, distortion_coeffs

def draw_axes(img, origin, imgpts):
    """
    Draw 3 coordinate axes on an image given an origin and projected axis endpoints.
    (This function is no longer used in the real-time display.)
    """
    imgpts = imgpts.astype(int)
    colors = [(0, 0, 255), (0, 255, 0), (255, 0, 0)]
    for i, pt in enumerate(imgpts):
        pt_tuple = tuple(pt.ravel())
        img = cv2.line(img, origin, pt_tuple, colors[i], 3)
    return img

def reorder_face_points4(face_points):
    """
    Reorder the four points of a face so that the first point is the reference,
    and the others are ordered based on distance from the reference.
    """
    p0 = face_points[0]
    others = face_points[1:]
    dist_list = []
    for i, pt in enumerate(others):
        d = np.linalg.norm(pt - p0)
        dist_list.append((d, i))
    dist_list.sort(key=lambda x: x[0])
    idx1 = dist_list[0][1]
    idx2 = dist_list[1][1]
    idx3 = dist_list[2][1]
    p1 = others[idx1]
    p2 = others[idx2]
    p3 = others[idx3]
    new_points = np.array([p0, p1, p2, p3], dtype=float)
    return new_points

def face_local_uv(intersection_point, face_points, invert_x=False, invert_y=False):
    """
    Compute local UV coordinates of an intersection point on a face.
    """
    new_pts = reorder_face_points4(face_points)
    p0, p1, p2, p3 = new_pts
    u_axis = p1 - p0
    v_axis = p2 - p0
    u_length = np.linalg.norm(u_axis)
    v_length = np.linalg.norm(v_axis)
    u_axis_norm = u_axis / (u_length + 1e-12)
    v_axis_norm = v_axis / (v_length + 1e-12)
    local_vec = intersection_point - p0
    u_local = np.dot(local_vec, u_axis_norm)
    v_local = np.dot(local_vec, v_axis_norm)
    if invert_x:
        u_local = u_length - u_local
    if invert_y:
        v_local = v_length - v_local
    return (u_local, v_local, u_length, v_length)

def intersect_line_with_plane(p0, dir_vec, plane_points):
    """
    Compute the intersection of a line (origin p0 and direction dir_vec) with a plane defined by four points.
    """
    p1, p2, p3, p4 = plane_points
    normal = np.cross(p2 - p1, p3 - p1)
    nrm = np.linalg.norm(normal)
    if nrm < 1e-12:
        return None
    normal /= nrm
    denom = np.dot(normal, dir_vec)
    if abs(denom) < 1e-9:
        return None
    d = np.dot(normal, p1 - p0) / denom
    if d < 0:
        return None
    return d

def area_of_triangle(a, b, c):
    """
    Compute the area of a triangle given by three points.
    """
    return 0.5 * np.linalg.norm(np.cross(b - a, c - a))

def point_in_polygon(point, polygon):
    """
    Check if a point lies within a convex quadrilateral (polygon).
    """
    p1, p2, p3, p4 = polygon
    area_orig = area_of_triangle(p1, p2, p3) + area_of_triangle(p1, p3, p4)
    area_test = (area_of_triangle(point, p1, p2) +
                 area_of_triangle(point, p2, p3) +
                 area_of_triangle(point, p3, p4) +
                 area_of_triangle(point, p4, p1))
    return np.isclose(area_orig, area_test, atol=1e-9)

def intersect_line_with_polygon(p0, dir_vec, polygon):
    """
    Compute the intersection of a line with a polygon. Returns the intersection point if inside.
    """
    t = intersect_line_with_plane(p0, dir_vec, polygon)
    if t is None:
        return None
    intersect_point = p0 + t * dir_vec
    if point_in_polygon(intersect_point, polygon):
        return intersect_point
    return None

# Note: Intersection calculation uses cubes for computation (in meters)
def find_frontmost_intersection(p0, dir_vec, all_cubes):
    """
    Find the frontmost intersection point between a line and a set of cubes.
    """
    closest_dist = float('inf')
    closest_hit = None
    for cube_data, cube_faces in all_cubes:
        for (face_idx, face_points) in cube_faces:
            ipt = intersect_line_with_polygon(p0, dir_vec, face_points)
            if ipt is not None:
                dist = np.linalg.norm(ipt - p0)
                if dist < closest_dist:
                    closest_dist = dist
                    closest_hit = (cube_data, face_idx, face_points, ipt)
    return closest_hit

def stable_nose_direction(R_local, old_nose_dir, nose_world_3d, camera_world_position, angle_threshold_deg=60.0):
    """
    Stabilize the nose direction vector to avoid abrupt changes.
    """
    new_z = R_local[:, 2]
    to_camera = (camera_world_position - nose_world_3d).ravel()
    if old_nose_dir is None:
        if np.dot(new_z, to_camera) < 0:
            R_local[:, 2] = -new_z
        return R_local
    dot_ = np.dot(old_nose_dir, new_z)
    dot_clamped = np.clip(dot_, -1.0, 1.0)
    angle_deg = np.degrees(np.arccos(dot_clamped))
    if angle_deg < angle_threshold_deg:
        return R_local
    if np.dot(new_z, to_camera) < 0:
        R_local[:, 2] = -new_z
    return R_local

def clear_all_3d_objects(ax, track_draw_data, intersection_draw_data):
    """
    Remove all 3D objects from the matplotlib axis.
    """
    for tid, tdd in track_draw_data.items():
        scat_obj = tdd["scatter"]
        line_obj = tdd["nose_line"]
        head_obj = tdd["nose_head"]
        if scat_obj is not None:
            scat_obj.remove()
        if line_obj is not None:
            line_obj.remove()
        if head_obj is not None:
            head_obj.remove()
        if tid in intersection_draw_data and intersection_draw_data[tid] is not None:
            intersection_draw_data[tid].remove()
    track_draw_data.clear()
    intersection_draw_data.clear()

def show_log_window():
    """
    Display the log messages in a separate OpenCV window.
    """
    width = 800
    height = 400
    log_canvas = np.zeros((height, width, 3), dtype=np.uint8)
    x0, y0 = 10, 20
    dy = 18
    color_text = (255, 255, 255)
    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.5
    thickness = 1
    lines_snapshot = list(log_deque)
    for i, line in enumerate(lines_snapshot):
        y = y0 + i * dy
        if y + 5 > height:
            break
        cv2.putText(log_canvas, line, (x0, y), font, font_scale, color_text, thickness, cv2.LINE_AA)
    cv2.imshow("Log Window", log_canvas)

def define_local_coordinate_system(points_world_dict, stable_points_set, initial_R_local=None):
    """
    Define a local coordinate system based on stable world points.
    """
    stable_points = []
    for sp in stable_points_set:
        if sp not in points_world_dict:
            return None, None
        stable_points.append(points_world_dict[sp].ravel())
    stable_points = np.array(stable_points)
    centroid = np.mean(stable_points, axis=0, keepdims=True)
    centered = stable_points - centroid
    U, S, Vt = np.linalg.svd(centered, full_matrices=False)
    x_axis = Vt[0, :].reshape(3, 1)
    y_axis = Vt[1, :].reshape(3, 1)
    z_axis = Vt[2, :].reshape(3, 1)
    x_axis /= np.linalg.norm(x_axis)
    y_axis /= np.linalg.norm(y_axis)
    z_axis /= np.linalg.norm(z_axis)
    R_local = np.hstack([x_axis, y_axis, z_axis])
    if np.linalg.det(R_local) < 0:
        R_local[:, 2] = -R_local[:, 2]
    if initial_R_local is not None:
        old_z = initial_R_local[:, 2]
        new_z = R_local[:, 2]
        if np.dot(old_z, new_z) < 0:
            R_local = -R_local
    origin = centroid.T
    return origin, R_local

##########################################################
# Initialize MediaPipe drawing tools for face mesh landmarks
##########################################################
mp_drawing = mp.solutions.drawing_utils
mp_drawing_styles = mp.solutions.drawing_styles

##########################################################
# Main Program
##########################################################
def main():
    logging.info("Program started.")
    # Start thread to flush log records to Excel
    flush_thread = threading.Thread(target=flush_thread_func, args=(XLSX_PATH,), daemon=True)
    flush_thread.start()

    # Define reference points for nose stabilization (in meters)
    reference_points = np.array([
        [0.0, 0.0, 0.0],
        [-0.03, 0.02, 0.0],
        [-0.02, 0.02, 0.0],
        [0.03, 0.02, 0.0],
        [0.02, 0.02, 0.0],
        [-0.02, -0.02, 0.0],
        [0.02, -0.02, 0.0],
        [0.0, -0.03, 0.0]
    ], dtype=np.float32).T

    # Get the existing maximum Face_ID from Excel to continue numbering
    existing_max_id = get_max_face_id_from_xlsm(XLSX_PATH)
    if existing_max_id < 0:
        existing_max_id = -1
    KalmanBoxTracker.count = existing_max_id + 1
    logging.info(f"Loaded existing max Face_ID = {existing_max_id}, next new ID = {KalmanBoxTracker.count}")

    # Parse command-line arguments
    parser = argparse.ArgumentParser()
    parser.add_argument("--enable_log", action="store_true", default=True,
                        help="Log intersection data into Excel (default: True)")
    parser.add_argument("--enable_3d_plot", action="store_true", default=False,
                        help="Display a 3D plot (default: False)")
    parser.add_argument("--enable_display", action="store_true", default=False,
                        help="Show real-time video in an OpenCV window (default: False)")
    parser.add_argument("--rotate180", action="store_false", default=True,
                        help="Disable 180 degree flip of camera image if specified")
    args = parser.parse_args()

    runtime_flip180 = args.rotate180

    # Load camera calibration files and intrinsic parameters
    rvec_path = os.path.join(IMAGE_FOLDER, RVEC_PATH)
    tvec_path = os.path.join(IMAGE_FOLDER, TVEC_PATH)
    camera_matrix_path = os.path.join(IMAGE_FOLDER, CAMERA_MATRIX_PATH)
    dist_coeffs_path = os.path.join(IMAGE_FOLDER, DIST_COEFFS_PATH)
    if not os.path.exists(camera_matrix_path):
        logging.error(f"Camera intrinsic matrix file not found: {camera_matrix_path}")
        raise FileNotFoundError(f"Camera intrinsic matrix file not found: {camera_matrix_path}")

    camera_matrix = np.load(camera_matrix_path).astype(np.float32)
    dist_coeffs = load_distortion_coeffs(dist_coeffs_path)
    rvec, tvec = load_vectors(rvec_path, tvec_path)

    # Convert rotation vector to rotation matrix
    R_loaded, _ = cv2.Rodrigues(rvec)
    R_inv = R_loaded.T
    t_wc = -R_loaded.T @ tvec

    # Initialize the RealSense pipeline and get camera parameters
    pipeline, depth_scale, intrinsics_realsense, distortion_coeffs_realsense = initialize_realsense()
    fx_realsense = intrinsics_realsense.fx
    fy_realsense = intrinsics_realsense.fy
    cx_realsense = intrinsics_realsense.ppx
    cy_realsense = intrinsics_realsense.ppy

    # Configure RealSense filters for depth image processing
    spatial = rs.spatial_filter()
    spatial.set_option(rs.option.holes_fill, 3)
    spatial.set_option(rs.option.filter_smooth_alpha, 0.5)
    spatial.set_option(rs.option.filter_smooth_delta, 20)
    spatial.set_option(rs.option.filter_magnitude, 2)

    temporal = rs.temporal_filter()
    temporal.set_option(rs.option.filter_smooth_alpha, 0.4)
    temporal.set_option(rs.option.filter_smooth_delta, 20)

    hole_filling = rs.hole_filling_filter(2)
    align_to = rs.stream.color
    align = rs.align(align_to)

    # Initialize MediaPipe Face Mesh detector
    mp_face_mesh = mp.solutions.face_mesh
    face_mesh = mp_face_mesh.FaceMesh(
        static_image_mode=False,
        max_num_faces=10,
        refine_landmarks=True,
        min_detection_confidence=MIN_DETECTION_CONFIDENCE,
        min_tracking_confidence=MIN_TRACKING_CONFIDENCE
    )
    logging.info("MediaPipe FaceMesh initialized.")

    # Initialize the 3D plot if enabled
    fig = None
    ax = None
    if args.enable_3d_plot:
        fig = plt.figure(figsize=(8, 6))
        ax = fig.add_subplot(111, projection='3d')
        ax.set_proj_type('persp')
        # Set axis limits in feet (display units)
        ax.set_xlim(X_LIMITS)
        ax.set_ylim(Y_LIMITS)
        ax.set_zlim(Z_LIMITS)
        ax.set_xlabel('X (ft)')
        ax.set_ylabel('Y (ft)')
        ax.set_zlabel('Z (ft)')
        ax.set_title('Multi-Face Nose Vector in 3D (ft)')
        ax.grid(True)
        ax.invert_zaxis()
        ax.invert_yaxis()
        plt.show(block=False)

    # Define conversion factors: original cube data is in feet.
    # Convert feet to meters for computations, and back to feet for display.
    FT_TO_M = 0.3048
    M_TO_FT = 1 / FT_TO_M

    # Build cube data for computations (in meters) and for display (in feet)
    all_cubes_calc = []  # For computation (meters)
    all_cubes_plot = []  # For display (feet)
    for cube in CUBES_INFO:
        face_points_ft = np.array(cube["face_points"], dtype=float)  # in feet
        depth_point_ft = np.array(cube["depth_point"], dtype=float)  # in feet
        # Convert to meters for computation
        face_points_m = face_points_ft * FT_TO_M
        depth_point_m = depth_point_ft * FT_TO_M

        try:
            faces_m = plot_cube_with_index(None, face_points_m, depth_point_m, color='gray', alpha=0.0)
        except Exception as e:
            logging.error(f"Error in cube computation: {e}")
            faces_m = []
        all_cubes_calc.append((cube, faces_m))

        if ax is not None:
            # For display, use original feet values and cube color from CUBES_INFO
            faces_ft = plot_cube_with_index(ax, face_points_ft, depth_point_ft, color=cube["color"], alpha=0.6)
            front_face = faces_ft[0][1]
            plot_face_points_with_colors(ax, front_face, colors=['red', 'green', 'blue', 'yellow'], s=80)
            all_cubes_plot.append((cube, faces_ft))
    if ax is not None:
        legend_patches = [mpatches.Patch(color=cube["color"], label=cube["name"]) for cube, _ in all_cubes_plot]
        ax.legend(handles=legend_patches, loc='upper right')

    # Dictionaries to store drawing objects for each tracked face
    track_draw_data = {}
    intersection_draw_data = {}
    mot_tracker = Sort(max_age=10, min_hits=2)

    # Variables for FPS calculation
    fps_frame_count = 0
    fps_start_time = time.time()
    fps = 0.0

    logging.info("Entering main loop.")
    try:
        while True:
            # Retrieve frames from RealSense
            frames = pipeline.wait_for_frames()
            if not frames:
                logging.warning("No frames retrieved, continue to next loop.")
                show_log_window()
                cv2.waitKey(1)
                time.sleep(FRAME_INTERVAL)
                continue

            color_frame_original = frames.get_color_frame()
            if not color_frame_original:
                logging.warning("No color frame retrieved, skipping iteration.")
                show_log_window()
                cv2.waitKey(1)
                time.sleep(FRAME_INTERVAL)
                continue

            # Convert the color frame to a NumPy array
            color_image_original = np.asanyarray(color_frame_original.get_data())
            if runtime_flip180:
                color_image_original = cv2.rotate(color_image_original, cv2.ROTATE_180)
            orig_height, orig_width = color_image_original.shape[:2]

            # Crop the image if needed
            if CROP_WIDTH > 0 and CROP_HEIGHT > 0:
                crop_width = CROP_WIDTH
                crop_height = CROP_HEIGHT
                crop_x = (orig_width - crop_width) // 2
                crop_y = (orig_height - crop_height) // 2
                color_image_cropped = color_image_original[crop_y:crop_y + crop_height,
                                      crop_x:crop_x + crop_width].copy()
                detection_input = color_image_cropped.copy()
            else:
                detection_input = color_image_original.copy()
                crop_x = 0
                crop_y = 0
                crop_width = orig_width
                crop_height = orig_height

            # Adjust the camera intrinsic matrix for cropping
            new_camera_matrix = camera_matrix.copy()
            if CROP_WIDTH > 0 and CROP_HEIGHT > 0:
                new_camera_matrix[0, 2] -= crop_x
                new_camera_matrix[1, 2] -= crop_y

            # Convert BGR to RGB for MediaPipe processing
            detection_input_rgb = cv2.cvtColor(detection_input, cv2.COLOR_BGR2RGB)
            results = face_mesh.process(detection_input_rgb)

            # Create a copy of the image for visualization if enabled
            img_vis = detection_input.copy() if args.enable_display else None

            face_bboxes = []
            face_landmarks_dict = []
            if results and results.multi_face_landmarks:
                for fm_idx, face_landmarks in enumerate(results.multi_face_landmarks):
                    px = []
                    py = []
                    for lm in face_landmarks.landmark:
                        x = lm.x * crop_width
                        y = lm.y * crop_height
                        px.append(x)
                        py.append(y)
                    x1 = max(0, np.min(px))
                    y1 = max(0, np.min(py))
                    x2 = min(crop_width, np.max(px))
                    y2 = min(crop_height, np.max(py))
                    face_bboxes.append([x1, y1, x2, y2])
                    face_landmarks_dict.append(face_landmarks)

                # Draw face mesh landmarks on the display image if enabled
                if img_vis is not None:
                    for face_landmarks in results.multi_face_landmarks:
                        mp_drawing.draw_landmarks(
                            image=img_vis,
                            landmark_list=face_landmarks,
                            connections=mp_face_mesh.FACEMESH_TESSELATION,
                            landmark_drawing_spec=mp_drawing.DrawingSpec(color=(0, 255, 0), thickness=1, circle_radius=1),
                            connection_drawing_spec=mp_drawing_styles.get_default_face_mesh_tesselation_style()
                        )
            else:
                if ax is not None:
                    clear_all_3d_objects(ax, track_draw_data, intersection_draw_data)
                if img_vis is not None:
                    fps_frame_count += 1
                    elapsed_time = time.time() - fps_start_time
                    if elapsed_time >= 1.0:
                        fps = fps_frame_count / elapsed_time
                        fps_frame_count = 0
                        fps_start_time = time.time()
                    cv2.putText(img_vis, f"FPS: {fps:.2f}", (50, 50),
                                cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)
                    cv2.putText(img_vis, "Press 'r' to flip 180", (50, 90),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 200, 200), 2)
                    cv2.imshow("Cropped - Frontmost Intersection", img_vis)
                show_log_window()
                key = cv2.waitKey(1)
                if key & 0xFF == ord('r'):
                    runtime_flip180 = not runtime_flip180
                    logging.info(f"Toggled runtime_flip180 => {runtime_flip180}")
                if key & 0xFF in [ord('q'), 27]:
                    logging.info("User pressed q/ESC, exiting main loop.")
                    break
                if ax is not None:
                    fig.canvas.draw()
                    fig.canvas.flush_events()
                time.sleep(FRAME_INTERVAL)
                continue

            # Align frames and get the depth frame
            aligned_frames = align.process(frames)
            depth_frame = aligned_frames.get_depth_frame()
            color_frame_aligned = aligned_frames.get_color_frame()
            if not depth_frame or not color_frame_aligned:
                show_log_window()
                cv2.waitKey(1)
                time.sleep(FRAME_INTERVAL)
                continue

            # Convert aligned frames to NumPy arrays
            color_image_aligned = np.asanyarray(color_frame_aligned.get_data())
            if runtime_flip180:
                color_image_aligned = cv2.rotate(color_image_aligned, cv2.ROTATE_180)
            depth_image = np.asanyarray(depth_frame.get_data()) * depth_scale
            if runtime_flip180:
                depth_image = cv2.rotate(depth_image, cv2.ROTATE_180)

            # Crop the aligned images to the same region
            color_image = color_image_aligned[crop_y:crop_y + crop_height, crop_x:crop_x + crop_width].copy()
            depth_image = depth_image[crop_y:crop_y + crop_height, crop_x:crop_x + crop_width].copy()

            if args.enable_display:
                img_vis = color_image.copy()

            # Update the SORT tracker with the face bounding boxes
            trackers_result = mot_tracker.update(face_bboxes)
            current_active_ids = set()

            for trk_data in trackers_result:
                x1, y1, x2, y2, tid = trk_data
                current_active_ids.add(tid)
                if img_vis is not None:
                    cv2.rectangle(img_vis, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 2)
                    cv2.putText(img_vis, f"ID: {tid}", (int(x1), int(y1) - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

                # Find the best matching face bounding box using IOU
                best_iou = -1
                best_fm_idx = -1
                for fm_idx, fb in enumerate(face_bboxes):
                    iou_val = mot_tracker.iou([x1, y1, x2, y2], fb)
                    if iou_val > best_iou:
                        best_iou = iou_val
                        best_fm_idx = fm_idx
                if best_fm_idx == -1:
                    continue

                face_landmarks = face_landmarks_dict[best_fm_idx]
                from_landmark_idx = [idx for idx in REFERENCE_LANDMARK_INDICES if idx < len(face_landmarks.landmark)]
                landmarks_pixels = {}
                for idx_ in from_landmark_idx:
                    xx = int(face_landmarks.landmark[idx_].x * crop_width)
                    yy = int(face_landmarks.landmark[idx_].y * crop_height)
                    landmarks_pixels[idx_] = (xx, yy)

                detected_points_world = {}
                valid_idxs = list(landmarks_pixels.keys())
                if len(valid_idxs) > 0:
                    coords = np.array([landmarks_pixels[k_] for k_ in valid_idxs], dtype=np.float32)
                    xs = coords[:, 0].astype(int)
                    ys = coords[:, 1].astype(int)
                    valid_in_img = []
                    for i_ in range(len(xs)):
                        if (0 <= ys[i_] < depth_image.shape[0]) and (0 <= xs[i_] < depth_image.shape[1]):
                            valid_in_img.append(i_)
                    if len(valid_in_img) > 0:
                        xs_in = xs[valid_in_img]
                        ys_in = ys[valid_in_img]
                        sub_valid_idxs = [valid_idxs[i_] for i_ in valid_in_img]
                        depths = depth_image[ys_in, xs_in]
                        valid_mask = depths > 0
                        if img_vis is not None:
                            for i_ in range(len(xs)):
                                ccenter = (xs[i_], ys[i_])
                                ccolor = (0, 0, 255)
                                if i_ in valid_in_img:
                                    idxv = valid_in_img.index(i_)
                                    if idxv < len(valid_mask) and valid_mask[idxv]:
                                        ccolor = (0, 255, 0)
                                cv2.circle(img_vis, ccenter, 3, ccolor, -1)
                        xs_in = xs_in[valid_mask]
                        ys_in = ys_in[valid_mask]
                        sub_valid_idxs = np.array(sub_valid_idxs)[valid_mask]
                        depths = depths[valid_mask]
                        if len(xs_in) > 0:
                            # Convert pixel coordinates to camera coordinates (in meters)
                            X = (xs_in - cx_realsense + crop_x) * depths / fx_realsense
                            Y = (ys_in - cy_realsense + crop_y) * depths / fy_realsense
                            Z = depths
                            camera_points = np.stack([X, Y, Z], axis=1)
                            world_arr = (R_inv @ camera_points.T + t_wc).T
                            for i_, k_ in enumerate(sub_valid_idxs):
                                detected_points_world[k_] = world_arr[i_].reshape(3, 1)

                if tid not in track_draw_data:
                    # Create drawing objects for the nose point and nose vector.
                    # Nose point is displayed in red and intersection point in lime.
                    scobj = None
                    lineobj = None
                    headobj = None
                    if ax is not None:
                        scobj = ax.scatter([], [], [], s=50, color='red')  # Nose origin (red)
                        lineobj = ax.plot([], [], [], linewidth=2, color='blue')[0]  # Nose vector (blue)
                        headobj = Poly3DCollection([], facecolors='red', edgecolors='none')
                        ax.add_collection3d(headobj)
                    track_draw_data[tid] = {
                        "world_points_history": {lid: [] for lid in REFERENCE_LANDMARK_INDICES},
                        "R_local_history": [],
                        "initial_R_local": None,
                        "frames_after_init": 0,
                        "scatter": scobj,
                        "nose_line": lineobj,
                        "nose_head": headobj,
                        "last_nose_dir": None
                    }
                    if ax is not None:
                        sc_int = ax.scatter([], [], [], c='lime', marker='o', s=60)  # Intersection point (lime)
                        intersection_draw_data[tid] = sc_int
                    else:
                        intersection_draw_data[tid] = None

                tdd = track_draw_data[tid]
                for k_ in detected_points_world:
                    tdd["world_points_history"][k_].append(detected_points_world[k_])

                if len(detected_points_world) >= 3:
                    def compute_variances(point_history):
                        variances = {}
                        for idx, hist in point_history.items():
                            if len(hist) < 2:
                                variances[idx] = 1.0
                            else:
                                arr = np.hstack(hist)
                                var = np.var(arr, axis=1).mean()
                                variances[idx] = var
                        return variances

                    variances = compute_variances(tdd["world_points_history"])
                    used_vars = [variances[kx] for kx in detected_points_world.keys()]
                    used_vars = np.array(used_vars)
                    weights = 1.0 / (used_vars + 1e-6)
                    weights /= (np.sum(weights) + 1e-12)
                    sub_idx_list = list(detected_points_world.keys())
                    det_points = np.hstack([detected_points_world[kx] for kx in sub_idx_list])
                    used_ref_indices = [REFERENCE_LANDMARK_INDICES.index(kx) for kx in sub_idx_list]
                    ref_subset = reference_points[:, used_ref_indices]
                    (R_kabsch, t_kabsch), inliers = ransac_weighted_kabsch(
                        ref_subset, det_points, weights,
                        threshold=RANSAC_THRESHOLD,
                        max_iterations=RANSAC_MAX_ITERATIONS
                    )
                    origin, R_local = define_local_coordinate_system(
                        detected_points_world, STABLE_POINTS_SET, tdd["initial_R_local"]
                    )
                    if R_local is not None:
                        # Compute nose origin in meters
                        nose_world_3d = origin.ravel()
                        R_local = stable_nose_direction(
                            R_local,
                            tdd["last_nose_dir"],
                            nose_world_3d,
                            t_wc.ravel(),
                            angle_threshold_deg=60.0
                        )
                        tdd["last_nose_dir"] = R_local[:, 2].copy()
                        tdd["R_local_history"].append(R_local)
                        if len(tdd["R_local_history"]) >= SMOOTH_WINDOW_SIZE:
                            R_stack = np.stack(tdd["R_local_history"][-SMOOTH_WINDOW_SIZE:], axis=0)
                            R_mean = np.mean(R_stack, axis=0)
                            U_, _, Vt_ = np.linalg.svd(R_mean)
                            R_ortho = U_ @ Vt_
                            R_local = R_ortho
                        tdd["frames_after_init"] += 1
                        if (tdd["initial_R_local"] is None) and (tdd["frames_after_init"] > INITIALIZATION_FRAMES):
                            tdd["initial_R_local"] = R_local.copy()
                        # Compute the forward point (nose tip extended) in meters
                        nose_forward_3d = origin + R_local[:, 2].reshape(3, 1) * FORWARD_LENGTH_3D
                        forward_point_3d = nose_forward_3d.ravel()

                        ### Update the 3D plot (display version): Convert points from meters to feet
                        nose_world_ft = nose_world_3d * M_TO_FT
                        forward_point_ft = forward_point_3d * M_TO_FT
                        if ax is not None:
                            # Update nose origin scatter (red)
                            scobj = tdd["scatter"]
                            if scobj is not None:
                                scobj._offsets3d = ([nose_world_ft[0]], [nose_world_ft[1]], [nose_world_ft[2]])
                            # Update the nose direction line (blue)
                            nose_line = tdd["nose_line"]
                            if nose_line is not None:
                                xs = [nose_world_ft[0], forward_point_ft[0]]
                                ys = [nose_world_ft[1], forward_point_ft[1]]
                                zs = [nose_world_ft[2], forward_point_ft[2]]
                                nose_line.set_data_3d(xs, ys, zs)
                        ### Update the intersection point scatter (lime)
                        sc_int = intersection_draw_data[tid]
                        closest_hit = find_frontmost_intersection(
                            nose_world_3d,
                            forward_point_3d - nose_world_3d,
                            all_cubes_calc
                        )
                        if closest_hit is not None:
                            cube_data, face_idx, face_points, ipt_3d = closest_hit
                            # Log if the face index matches the record_face, otherwise only update display
                            if face_idx == cube_data.get("record_face", 0):
                                invert_x = cube_data.get("invert_x", False)
                                invert_y = cube_data.get("invert_y", False)
                                u_local, v_local, u_len, v_len = face_local_uv(
                                    ipt_3d,
                                    face_points,
                                    invert_x=invert_x,
                                    invert_y=invert_y
                                )
                                plane_name = cube_data["name"]
                                now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                logging.info(f"[HIT] ID={tid}, plane={plane_name}, face_idx={face_idx}, "
                                             f"hit=({ipt_3d[0]:.3f},{ipt_3d[1]:.3f},{ipt_3d[2]:.3f}), "
                                             f"u_local={u_local:.3f}/{u_len:.3f}, v_local={v_local:.3f}/{v_len:.3f}")
                                if args.enable_log:
                                    log_intersection_2d_async(tid, plane_name, u_local, v_local, now_str)
                            else:
                                logging.info(f"[BLOCK] ID={tid}, face_idx={face_idx} does not match record_face.")
                            if ax is not None and sc_int is not None:
                                ipt_ft = ipt_3d * M_TO_FT
                                sc_int._offsets3d = ([ipt_ft[0]], [ipt_ft[1]], [ipt_ft[2]])
                        else:
                            if ax is not None and sc_int is not None:
                                sc_int._offsets3d = ([], [], [])
                        ### Draw the nose vector in the 2D real-time display
                        if img_vis is not None:
                            nose_forward_2d = origin + R_local[:, 2].reshape(3, 1) * FORWARD_LENGTH_2D
                            nose_3d_pts = np.vstack([origin.T, nose_forward_2d.T])
                            nose_imgpts, _ = cv2.projectPoints(nose_3d_pts, rvec, tvec, new_camera_matrix, dist_coeffs)
                            nose_imgpts = nose_imgpts.reshape(-1, 2).astype(int)
                            start_pt_2d = tuple(nose_imgpts[0])
                            end_pt_2d = tuple(nose_imgpts[1])
                            cv2.arrowedLine(img_vis, start_pt_2d, end_pt_2d, (0, 255, 255), 3, tipLength=0.2)
                            cv2.circle(img_vis, end_pt_2d, 5, (0, 255, 255), -1)
            for old_tid in list(track_draw_data.keys()):
                if old_tid not in current_active_ids:
                    if ax is not None:
                        scat_obj = track_draw_data[old_tid]["scatter"]
                        line_obj = track_draw_data[old_tid]["nose_line"]
                        head_obj = track_draw_data[old_tid]["nose_head"]
                        if scat_obj is not None:
                            scat_obj.remove()
                        if line_obj is not None:
                            line_obj.remove()
                        if head_obj is not None:
                            head_obj.remove()
                        if old_tid in intersection_draw_data and intersection_draw_data[old_tid] is not None:
                            intersection_draw_data[old_tid].remove()
                            del intersection_draw_data[old_tid]
                    del track_draw_data[old_tid]
            if img_vis is not None:
                # Note: The following block that draws the coordinate axes on the 2D display has been removed.
                # This prevents the real-time display from showing the three axes representing the real-world origin.
                fps_frame_count += 1
                elapsed_time = time.time() - fps_start_time
                if elapsed_time >= 1.0:
                    fps = fps_frame_count / elapsed_time
                    fps_frame_count = 0
                    fps_start_time = time.time()
                cv2.putText(img_vis, f"FPS: {fps:.2f}", (50, 50),
                            cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)
                cv2.putText(img_vis, "Press 'r' to flip 180", (50, 90),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 200, 200), 2)
                cv2.imshow("Cropped - Frontmost Intersection", img_vis)
            show_log_window()
            key = cv2.waitKey(1)
            if key & 0xFF == ord('r'):
                runtime_flip180 = not runtime_flip180
                logging.info(f"Toggled runtime_flip180 => {runtime_flip180}")
            if key & 0xFF in [ord('q'), 27]:
                logging.info("User pressed q/ESC, exiting main loop.")
                break
            if ax is not None:
                fig.canvas.draw()
                fig.canvas.flush_events()
            time.sleep(FRAME_INTERVAL)
    except KeyboardInterrupt:
        logging.info("User interrupted by Ctrl+C.")
    finally:
        logging.info("Exiting, start resource cleanup...")
        face_mesh.close()
        pipeline.stop()
        cv2.destroyAllWindows()
        if fig is not None:
            plt.close(fig)
        stop_event.set()
        flush_thread.join()
        logging.info("Program ended.")


if __name__ == "__main__":
    main()
